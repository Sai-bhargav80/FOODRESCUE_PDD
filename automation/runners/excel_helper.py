import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def style_sheet(ws):
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Dark slate
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    # Format header row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[1].height = 28
    
    # Auto column width
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val = str(cell.value or '')
            if '\n' in val:
                val = max(val.split('\n'), key=len)
            max_len = max(max_len, len(val))
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

def generate_report(output_path, title, test_cases):
    wb = openpyxl.Workbook()
    
    passed_cases = [tc for tc in test_cases if tc["status"] == "Passed"]
    failed_cases = [tc for tc in test_cases if tc["status"] == "Failed"]
    skipped_cases = [tc for tc in test_cases if tc["status"] == "Skipped"]
    
    # Sheet 1: Executed Test Cases
    ws_all = wb.active
    ws_all.title = "Executed Test Cases"
    ws_all.append(["Test ID", "Module", "Test Name", "Status", "Execution Time (s)", "Priority"])
    for tc in test_cases:
        ws_all.append([tc["id"], tc["module"], tc["name"], tc["status"], tc["duration"], tc["priority"]])
    style_sheet(ws_all)
    
    # Sheet 2: Passed Tests
    ws_pass = wb.create_sheet(title="Passed Tests")
    ws_pass.append(["Test ID", "Module", "Test Name", "Status", "Execution Time (s)", "Priority"])
    for tc in passed_cases:
        ws_pass.append([tc["id"], tc["module"], tc["name"], tc["status"], tc["duration"], tc["priority"]])
    style_sheet(ws_pass)
    
    # Sheet 3: Failed Tests
    ws_fail = wb.create_sheet(title="Failed Tests")
    ws_fail.append(["Test ID", "Module", "Test Name", "Status", "Execution Time (s)", "Priority", "Failure Reason"])
    for tc in failed_cases:
        ws_fail.append([tc["id"], tc["module"], tc["name"], tc["status"], tc["duration"], tc["priority"], tc.get("reason", "")])
    style_sheet(ws_fail)
    
    # Sheet 4: Skipped Tests
    ws_skip = wb.create_sheet(title="Skipped Tests")
    ws_skip.append(["Test ID", "Module", "Test Name", "Status", "Execution Time (s)", "Priority"])
    for tc in skipped_cases:
        ws_skip.append([tc["id"], tc["module"], tc["name"], tc["status"], tc["duration"], tc["priority"]])
    style_sheet(ws_skip)
    
    # Sheet 5: Execution Metrics
    ws_metrics = wb.create_sheet(title="Execution Metrics")
    ws_metrics.append(["Metric", "Value"])
    ws_metrics.append(["Total Executed", len(test_cases)])
    ws_metrics.append(["Passed", len(passed_cases)])
    ws_metrics.append(["Failed", len(failed_cases)])
    ws_metrics.append(["Skipped", len(skipped_cases)])
    pass_pct = (len(passed_cases) / len(test_cases) * 100) if test_cases else 0.0
    ws_metrics.append(["Pass Percentage (%)", f"{pass_pct:.2f}%"])
    total_dur = sum(tc["duration"] for tc in test_cases)
    ws_metrics.append(["Total Duration (s)", f"{total_dur:.2f}s"])
    style_sheet(ws_metrics)
    
    # Sheet 6: Defect Summary
    ws_defects = wb.create_sheet(title="Defect Summary")
    ws_defects.append(["Defect ID", "Test Case ID", "Module", "Severity", "Description", "Status"])
    for idx, tc in enumerate(failed_cases, start=1):
        ws_defects.append([f"DEFECT_{idx:03d}", tc["id"], tc["module"], tc["priority"], tc.get("reason", "Unknown assertion error"), "New"])
    style_sheet(ws_defects)
    
    # Save workbook
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
